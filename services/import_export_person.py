# services/import_export_person.py
# 人员专属导入导出逻辑（生产级终极版 - 2026-02-09）
# 更新：字段全面同步最新 schema，支持所有新字段（relationship、household_number、is_key_person 等）
# 修复：界面假成功问题 → 使用 create_person 单条插入，确保真实写入数据库
# 优化：导入支持更多列名变体、布尔字段更宽松、身份证重复友好提示

import os
import pandas as pd
from datetime import datetime
from flask import current_app
from flask_login import current_user
from repositories.person_repo import get_all_people_for_export, create_person
from repositories.building_repo import get_building_by_name_or_address
from permissions import check_user_grid_permission, get_user_grid_ids
from repositories.grid_repo import get_grid_by_id
from utils import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from werkzeug.utils import secure_filename


# 布尔宽松映射（支持更多表达方式）
def str_to_bool(val) -> int:
    if pd.isna(val):
        return 0
    val = str(val).strip().lower()
    return 1 if val in ['1', '是', 'true', 'yes', 'y', '有', '重点', '是重点'] else 0


def export_person_to_excel(user) -> tuple[str, str]:
    """导出人员数据到 Excel（带两行注释 + 所有字段）"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exports_dir = current_app.config['EXPORTS_FOLDER']
    os.makedirs(exports_dir, exist_ok=True)

    user_grid_ids = get_user_grid_ids(user)
    raw_data = get_all_people_for_export(grid_ids=user_grid_ids if user_grid_ids else None)

    filename_prefix = "人员数据"
    if user_grid_ids and len(user_grid_ids) == 1:
        grid = get_grid_by_id(user_grid_ids[0])
        grid_name = grid['name'] if grid else "未知网格"
        filename_prefix += f"_{grid_name}"

    # 完整表头（与 schema 顺序一致）
    headers = [
        '姓名', '身份证号', '唯一标识', '护照/其他证件号码', '其他证件类型',
        '性别', '出生日期', '联系电话', '现住小区/建筑', '现住详细门牌',
        '所属网格', '与其他人员关系', '人员类型', '是否重点人员', '重点类别',
        '户籍小区/建筑', '户籍详细地址', '户编号', '户号', '户籍迁入日期',
        '是否人户分离', '实际居住地', '是否已迁出', '迁出日期', '迁往地',
        '是否已死亡', '死亡日期', '民族', '政治面貌', '婚姻状况',
        '文化程度', '工作/学习情况', '健康状况', '备注'
    ]

    # 对应注释（引导用户填写规范）
    comments = [
        '必填，真实姓名', '可选，18位身份证号（无证可留空）', '系统内部唯一标识（可选）', '护照或其他证件号码', '护照/军人证/港澳通行证等',
        '男/女（支持：男、M、1；女、F、0）', '格式：YYYYMMDD', '多个用;分隔，可选', '系统内现住建筑名称（必填）', '如1单元101室（必填）',
        '自动关联，无需填写', '如：户主、配偶、子女、父母、租户（可选）', '常住人口/流动人口', '是/否 或 1/0', '多个类别用,分隔，如独居老人,低保户',
        '本社区户籍建筑名称（可选）', '外地户籍填写完整地址', '家庭编号（如001、A001）', '户口本户号', '格式：YYYYMMDD',
        '是/否 或 1/0', '人户分离时的实际居住地址', '是/否 或 1/0', '格式：YYYYMMDD', '迁往省市区',
        '是/否 或 1/0', '格式：YYYYMMDD', '如汉族、回族', '如中共党员、群众', '未婚/已婚/离异/丧偶',
        '小学/初中/高中/本科等', '在职/在校/退休/无业等', '健康/良好/慢性病/残疾等', '其他补充信息'
    ]

    processed_data = []
    for item in raw_data:
        row = {
            '姓名': item.get('name', ''),
            '身份证号': item.get('id_card', ''),
            '唯一标识': item.get('unique_id', ''),
            '护照/其他证件号码': item.get('passport', ''),
            '其他证件类型': item.get('other_id_type', ''),
            '性别': item.get('gender', ''),
            '出生日期': item.get('birth_date', ''),
            '联系电话': item.get('phones', ''),
            '现住小区/建筑': item.get('living_building_name', ''),
            '现住详细门牌': item.get('address_detail', ''),
            '所属网格': item.get('grid_name') or '无网格',
            '与其他人员关系': item.get('relationship', ''),
            '人员类型': item.get('person_type', ''),
            '是否重点人员': '是' if item.get('is_key_person') else '否',
            '重点类别': item.get('key_categories', ''),
            '户籍小区/建筑': item.get('household_building_name', '') if 'household_building_name' in item else '',
            '户籍详细地址': item.get('household_address', ''),
            '户编号': item.get('family_id', ''),
            '户号': item.get('household_number', ''),
            '户籍迁入日期': item.get('household_entry_date', ''),
            '是否人户分离': '是' if item.get('is_separated') else '否',
            '实际居住地': item.get('current_residence', ''),
            '是否已迁出': '是' if item.get('is_migrated_out') else '否',
            '迁出日期': item.get('household_exit_date', ''),
            '迁往地': item.get('migration_destination', ''),
            '是否已死亡': '是' if item.get('is_deceased') else '否',
            '死亡日期': item.get('death_date', ''),
            '民族': item.get('nationality', ''),
            '政治面貌': item.get('political_status', ''),
            '婚姻状况': item.get('marital_status', ''),
            '文化程度': item.get('education', ''),
            '工作/学习情况': item.get('work_study', ''),
            '健康状况': item.get('health', ''),
            '备注': item.get('notes', '')
        }
        processed_data.append(row)

    if not processed_data:
        processed_data = [dict.fromkeys(headers, '')]

    df = pd.DataFrame(processed_data)

    filename = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = os.path.join(exports_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = '人员数据'

    # 第一行：表头
    ws.append(headers)
    # 第二行：注释
    ws.append(comments)

    # 数据从第三行开始
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # 样式
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_align

    for cell in ws[2]:
        cell.alignment = left_align

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 4, 60)

    wb.save(file_path)
    logger.info(f"用户 {current_user.username} 导出人员数据: {filename}（共 {len(processed_data)} 条）")
    return file_path, filename


def import_person_from_excel(file, user) -> tuple[bool, str]:
    """生产级终极版：真实写入数据库 + 精准错误反馈 + 支持所有字段"""
    imports_folder = current_app.config['IMPORTS_FOLDER']
    temp_path = os.path.join(imports_folder, secure_filename(file.filename))
    file.save(temp_path)

    try:
        # 读取 Excel，全部转为字符串，空值填充空字符串
        df = pd.read_excel(temp_path, dtype=str).fillna('')

        # 必填列检查（与前端模板一致）
        required_columns = ['姓名', '现住小区/建筑', '现住详细门牌']
        actual_columns = [str(col).strip() for col in df.columns]
        missing = [col for col in required_columns if col not in actual_columns]
        if missing:
            return False, f'导入失败：缺少必填列 {", ".join(missing)}。<br>请使用最新导出模板，确保包含这些列。'

        success_count = 0
        fail_reasons = []

        for idx, row in df.iterrows():
            name = str(row.get('姓名', '')).strip()
            if not name:
                fail_reasons.append(f"第 {idx+2} 行：姓名为空，跳过")
                continue

            # 建筑匹配（必填）
            living_building_name = str(row.get('现住小区/建筑', '')).strip()
            if not living_building_name:
                fail_reasons.append(f"第 {idx+2} 行：现住小区/建筑为空（{name}）")
                continue

            living_building = get_building_by_name_or_address(living_building_name)
            if not living_building:
                fail_reasons.append(f"第 {idx+2} 行：未找到现住建筑 '{living_building_name}'（{name}）")
                continue

            if not check_user_grid_permission(living_building['id']):
                fail_reasons.append(f"第 {idx+2} 行：无权操作该网格建筑 '{living_building_name}'（{name}）")
                continue

            # 门牌（必填）
            address_detail = str(row.get('现住详细门牌', '')).strip()
            if not address_detail:
                fail_reasons.append(f"第 {idx+2} 行：现住详细门牌为空（{name}）")
                continue

            # 户籍建筑（可选）
            household_building_name = str(row.get('户籍小区/建筑', '')).strip()
            household_building_id = None
            if household_building_name:
                household_building = get_building_by_name_or_address(household_building_name)
                if household_building:
                    household_building_id = household_building['id']

            # 智能性别映射
            raw_gender = str(row.get('性别', '')).strip()
            gender = None
            if raw_gender in ['男', '男性', 'M', '1', '男士']:
                gender = '男'
            elif raw_gender in ['女', '女性', 'F', '0', '女士']:
                gender = '女'

            # 构建导入记录（字段完整）
            record = {
                'name': name,
                'id_card': str(row.get('身份证号', '')).strip() or None,
                'unique_id': str(row.get('唯一标识', '')).strip() or None,
                'passport': str(row.get('护照/其他证件号码', '')).strip() or None,
                'other_id_type': str(row.get('其他证件类型', '')).strip() or None,
                'phones': str(row.get('联系电话', '')).strip() or None,
                'gender': gender,
                'birth_date': str(row.get('出生日期', '')).strip() or None,
                'person_type': str(row.get('人员类型', '')).strip() or '常住人口',
                'relationship': str(row.get('与其他人员关系', '')).strip() or None,
                'living_building_id': living_building['id'],
                'address_detail': address_detail,
                'household_building_id': household_building_id,
                'household_address': str(row.get('户籍详细地址', '')).strip() or None,
                'family_id': str(row.get('户编号', '')).strip() or None,
                'household_number': str(row.get('户号', '')).strip() or None,
                'household_entry_date': str(row.get('户籍迁入日期', '')).strip() or None,
                'is_separated': str_to_bool(row.get('是否人户分离', '')),
                'current_residence': str(row.get('实际居住地', '')).strip() or None,
                'is_migrated_out': str_to_bool(row.get('是否已迁出', '')),
                'household_exit_date': str(row.get('迁出日期', '')).strip() or None,
                'migration_destination': str(row.get('迁往地', '')).strip() or None,
                'is_deceased': str_to_bool(row.get('是否已死亡', '')),
                'death_date': str(row.get('死亡日期', '')).strip() or None,
                'nationality': str(row.get('民族', '')).strip() or None,
                'political_status': str(row.get('政治面貌', '')).strip() or None,
                'marital_status': str(row.get('婚姻状况', '')).strip() or None,
                'education': str(row.get('文化程度', '')).strip() or None,
                'work_study': str(row.get('工作/学习情况', '')).strip() or None,
                'health': str(row.get('健康状况', '')).strip() or None,
                'notes': str(row.get('备注', '')).strip() or None,
                'is_key_person': str_to_bool(row.get('是否重点人员', '')),
                'key_categories': str(row.get('重点类别', '')).strip() or None,
            }

            try:
                # 单条真实插入
                create_person(**record)
                success_count += 1
            except Exception as row_e:
                error_msg = str(row_e)
                if "UNIQUE constraint failed: person.id_card" in error_msg:
                    fail_reasons.append(f"第 {idx+2} 行：身份证号 {record['id_card'] or '(空)'} 已存在（重复人员，自动跳过）")
                elif "NOT NULL constraint failed" in error_msg:
                    fail_reasons.append(f"第 {idx+2} 行：违反非空约束（{name}）")
                else:
                    fail_reasons.append(f"第 {idx+2} 行：{error_msg[:120]}...")

        fail_count = len(df) - success_count

        msg = f'人员导入完成：成功 {success_count} 条，失败 {fail_count} 条'
        if fail_reasons:
            sample = fail_reasons[:5]
            msg += f"<br>失败原因示例：{'；'.join(sample)}"
            if len(fail_reasons) > 5:
                msg += "（更多错误请查看服务器日志）"

        logger.info(f"用户 {user.username} 导入人员：真实成功 {success_count} 条，失败 {fail_count} 条")
        return True, msg

    except Exception as e:
        logger.error(f"导入人员整体异常: {e}")
        return False, f"导入失败：文件读取或处理异常（{str(e)[:100]}）"

    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.warning(f"清理临时导入文件失败: {temp_path} - {e}")
