# services/import_export_building.py
# 建筑专属导入导出逻辑（终极扩展版 - 2026-01-06）
# 新增：导入支持导出全部31个字段（宽松读取，无严格校验）
#       - 必填仍为：小区/建筑名称、类型、所属网格
#       - 其他字段可选存在即读取，缺失用默认值
#       - 数字字段宽松转int/float（失败用0或空）
#       - 布尔字段宽松映射（1/是/True → 1，其他 → 0）
#       - 类型映射保持原有逻辑

import os
import pandas as pd
from datetime import datetime
from flask import current_app
from flask_login import current_user
from repositories.building_repo import get_all_buildings_for_export, create_building
from repositories.grid_repo import get_all_grids, get_grid_by_id
from repositories.base import get_db_connection
from permissions import get_user_grid_ids
from utils import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from werkzeug.utils import secure_filename


# 建筑类型映射（保持原有）
BUILDING_TYPE_MAPPING = {
    '住宅小区': 'residential_complex',
    '商业建筑': 'commercial',
    '商业大厦': 'commercial',
    '大型出租房': 'large_rental',
    '公寓': 'large_rental',
    '私人住宅': 'private_residence',
    '其他': 'residential_complex',
}


# 宽松布尔转换
def str_to_bool(val: str) -> int:
    val = str(val).strip().lower()
    return 1 if val in ['1', '是', 'true', 'yes', 'y'] else 0


# 宽松数字转换
def str_to_int(val: str, default=0) -> int:
    try:
        return int(str(val).strip()) if str(val).strip() else default
    except:
        return default


def str_to_float(val: str, default=0.0) -> float:
    try:
        return float(str(val).strip()) if str(val).strip() else default
    except:
        return default


def export_building_to_excel(user) -> tuple[str, str]:
    # 【导出逻辑保持不变 - 已完美】
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exports_dir = current_app.config['EXPORTS_FOLDER']
    os.makedirs(exports_dir, exist_ok=True)

    user_grid_ids = get_user_grid_ids(user)
    raw_data = get_all_buildings_for_export(grid_ids=user_grid_ids if user_grid_ids else None)

    filename_prefix = "小区建筑数据"
    if user_grid_ids and len(user_grid_ids) == 1:
        grid = get_grid_by_id(user_grid_ids[0])
        grid_name = grid['name'] if grid else "未知网格"
        filename_prefix += f"_{grid_name}"

    headers = [
        '小区/建筑名称', '类型', '所属网格', '详细地址', '建成年份', '户数', '楼栋数', '约居住人数',
        '企业数', '底商数', '是否通燃气', '物业费标准', '电梯数', '室内车位数', '室外车位数',
        '安全负责人', '安全负责人电话', '纬度', '经度',
        '开发商', '施工单位', '物业公司', '物业联系电话', '备注',
        '业委会联系人', '业委会联系电话', '业主姓名', '业主电话',
        '房东姓名', '房东电话', '商业类型'
    ]

    comments = [
        '必填，在网格内唯一', '住宅小区/商业建筑/大型出租房/私人住宅', '必填，网格名称', '完整街道门牌号',
        '如2015', '总户数', '楼栋数量', '预估居住人数', '含居家办公企业', '底商店铺数量',
        '1=是，0=否', '如2.5元/㎡/月', '电梯总数', '室内停车位', '室外停车位',
        '安全负责人姓名', '联系电话', '可选，十进制纬度', '可选，十进制经度',
        '开发商名称', '施工单位名称', '物业公司全称', '物业联系电话', '其他补充信息',
        '住宅小区专用', '住宅小区专用', '私人住宅专用', '私人住宅专用',
        '大型出租房专用', '大型出租房专用', '商业建筑专用（如商场、写字楼）'
    ]

    type_display_map = {
        'residential_complex': '住宅小区',
        'commercial': '商业建筑',
        'large_rental': '大型出租房',
        'private_residence': '私人住宅'
    }

    processed_data = []
    for item in raw_data:
        row = {
            '小区/建筑名称': item.get('name', ''),
            '类型': type_display_map.get(item.get('type'), item.get('type', '未知')),
            '所属网格': item.get('grid_name') or '无网格',
            '详细地址': item.get('address', ''),
            '建成年份': item.get('build_year', ''),
            '户数': item.get('households', ''),
            '楼栋数': item.get('buildings_count', ''),
            '约居住人数': item.get('approx_residents', ''),
            '企业数': item.get('businesses_count', ''),
            '底商数': item.get('ground_floor_shops', ''),
            '是否通燃气': '是' if item.get('has_gas_pipeline') else '否',
            '物业费标准': item.get('property_fee', ''),
            '电梯数': item.get('elevators', ''),
            '室内车位数': item.get('indoor_parking', ''),
            '室外车位数': item.get('outdoor_parking', ''),
            '安全负责人': item.get('security_manager', ''),
            '安全负责人电话': item.get('security_manager_phone', ''),
            '纬度': item.get('latitude', ''),
            '经度': item.get('longitude', ''),
            '开发商': item.get('developer', ''),
            '施工单位': item.get('constructor', ''),
            '物业公司': item.get('property_management_company', ''),
            '物业联系电话': item.get('property_contact_phone', ''),
            '备注': item.get('notes', ''),
            '业委会联系人': item.get('owners_committee_contact', ''),
            '业委会联系电话': item.get('owners_committee_phone', ''),
            '业主姓名': item.get('owner_name', ''),
            '业主电话': item.get('owner_phone', ''),
            '房东姓名': item.get('landlord_name', ''),
            '房东电话': item.get('landlord_phone', ''),
            '商业类型': item.get('commercial_type', '')
        }
        processed_data.append(row)

    if not processed_data:
        processed_data = [dict.fromkeys(headers, '')]

    df = pd.DataFrame(processed_data)

    filename = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = os.path.join(exports_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Building'

    ws.append(headers)
    ws.append(comments)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_align
    for cell in ws[2]:
        cell.alignment = wrap_align

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value or "")) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(file_path)
    logger.info(f"用户 {current_user.username} 导出建筑数据: {filename}（共 {len(processed_data)} 条）")
    return file_path, filename


def import_building_from_excel(file, user) -> tuple[bool, str]:
    """扩展版：支持全部31字段导入（宽松、无校验）"""
    imports_folder = current_app.config['IMPORTS_FOLDER']
    temp_path = os.path.join(imports_folder, secure_filename(file.filename))
    file.save(temp_path)

    try:
        df = pd.read_excel(temp_path, dtype=str).fillna('')

        # 必填列检查（保持严格，避免垃圾数据）
        required_columns = ['小区/建筑名称', '类型', '所属网格']
        actual_columns = [str(col).strip() for col in df.columns]
        missing = [col for col in required_columns if col not in actual_columns]
        if missing:
            return False, f'建筑导入失败：缺少必填列 {", ".join(missing)}。<br>请下载最新模板并确保包含这些列。'

        all_grids = get_all_grids()
        grid_name_to_id = {g['name'].strip().lower(): g['id'] for g in all_grids}

        success_count = 0
        fail_reasons = []

        for idx, row in df.iterrows():
            name = str(row['小区/建筑名称']).strip()
            type_cn = str(row['类型']).strip()
            grid_name = str(row['所属网格']).strip()

            if not name:
                fail_reasons.append(f"第 {idx+2} 行：小区/建筑名称为空")
                continue
            if not type_cn:
                fail_reasons.append(f"第 {idx+2} 行：类型为空（{name}）")
                continue
            if not grid_name:
                fail_reasons.append(f"第 {idx+2} 行：所属网格为空（{name}）")
                continue

            # 类型映射（宽松模糊匹配）
            type_en = BUILDING_TYPE_MAPPING.get(type_cn)
            if not type_en:
                matched = False
                for key, val in BUILDING_TYPE_MAPPING.items():
                    if key in type_cn or type_cn in key:
                        type_en = val
                        matched = True
                        break
                if not matched:
                    fail_reasons.append(f"第 {idx+2} 行：未知建筑类型 '{type_cn}'（{name}），已默认住宅小区")
                    type_en = 'residential_complex'  # 默认 fallback

            # 网格匹配（原有智能匹配）
            grid_name_lower = grid_name.lower()
            grid_id = None
            if grid_name_lower in grid_name_to_id:
                grid_id = grid_name_to_id[grid_name_lower]
            else:
                matches = [gid for gname, gid in grid_name_to_id.items()
                           if grid_name_lower in gname or gname in grid_name_lower]
                if len(matches) == 1:
                    grid_id = matches[0]
                elif len(matches) > 1:
                    fail_reasons.append(f"第 {idx+2} 行：网格名称模糊匹配到多个（{name}）")
                    continue
                else:
                    fail_reasons.append(f"第 {idx+2} 行：未找到网格 '{grid_name}'（{name}）")
                    continue

            # 权限检查
            if not ('super_admin' in user.roles or 'community_admin' in user.roles or grid_id in user.managed_grids):
                fail_reasons.append(f"第 {idx+2} 行：无权操作网格 '{grid_name}'（{name}）")
                continue

            # 唯一性检查
            with get_db_connection() as conn:
                conflict = conn.execute(
                    "SELECT id FROM building WHERE name = ? AND grid_id = ? AND is_deleted = 0",
                    (name, grid_id)
                ).fetchone()
                if conflict:
                    fail_reasons.append(f"第 {idx+2} 行：该网格下已存在同名建筑 '{name}'")
                    continue

            # 宽松读取全部字段
            try:
                create_building(
                    name=name,
                    type_=type_en,
                    grid_id=grid_id,
                    address=str(row.get('详细地址', '')).strip(),
                    build_year=str_to_int(row.get('建成年份', '')),
                    households=str_to_int(row.get('户数', 0)),
                    buildings_count=str_to_int(row.get('楼栋数', 0)),
                    approx_residents=str_to_int(row.get('约居住人数', 0)),
                    businesses_count=str_to_int(row.get('企业数', 0)),
                    ground_floor_shops=str_to_int(row.get('底商数', 0)),
                    has_gas_pipeline=str_to_bool(row.get('是否通燃气', '')),
                    property_fee=str(row.get('物业费标准', '')).strip(),
                    elevators=str_to_int(row.get('电梯数', 0)),
                    indoor_parking=str_to_int(row.get('室内车位数', 0)),
                    outdoor_parking=str_to_int(row.get('室外车位数', 0)),
                    security_manager=str(row.get('安全负责人', '')).strip(),
                    security_manager_phone=str(row.get('安全负责人电话', '')).strip(),
                    latitude=str_to_float(row.get('纬度', '')),
                    longitude=str_to_float(row.get('经度', '')),
                    developer=str(row.get('开发商', '')).strip(),
                    constructor=str(row.get('施工单位', '')).strip(),
                    property_management_company=str(row.get('物业公司', '')).strip(),
                    property_contact_phone=str(row.get('物业联系电话', '')).strip(),
                    notes=str(row.get('备注', '')).strip(),
                    owners_committee_contact=str(row.get('业委会联系人', '')).strip(),
                    owners_committee_phone=str(row.get('业委会联系电话', '')).strip(),
                    owner_name=str(row.get('业主姓名', '')).strip(),
                    owner_phone=str(row.get('业主电话', '')).strip(),
                    landlord_name=str(row.get('房东姓名', '')).strip(),
                    landlord_phone=str(row.get('房东电话', '')).strip(),
                    commercial_type=str(row.get('商业类型', '')).strip()
                )
                success_count += 1
            except Exception as e:
                fail_reasons.append(f"第 {idx+2} 行：数据库错误 ({name}): {str(e)[:50]}")

        fail_count = len(df) - success_count
        msg = f'建筑导入完成：成功 {success_count} 条，失败 {fail_count} 条'
        if fail_reasons:
            sample = fail_reasons[:5]
            msg += f"<br>失败原因示例：{'；'.join(sample)}"
            if len(fail_reasons) > 5:
                msg += "（完整错误见服务器日志）"

        logger.info(f"用户 {user.username} 导入建筑：成功 {success_count}，失败 {fail_count}")
        return True, msg

    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.warning(f"清理临时文件失败: {temp_path} - {e}")
