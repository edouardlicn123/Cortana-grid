# services/import_export_service.py
# 导入导出核心业务逻辑（优化版 - 代码更简洁、可读性提升、健壮性更强，功能完全不变）

import os
import pandas as pd
from datetime import datetime
from flask import current_app
from flask_login import current_user
from werkzeug.utils import secure_filename
from repositories.building_repo import (
    get_all_buildings_for_export,
    get_building_by_name_or_address
)
from repositories.person_repo import get_all_people_for_export, bulk_insert_people
from repositories.grid_repo import get_grid_by_id
from permissions import check_user_grid_permission, get_user_grid_ids
from utils import logger


ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename: str) -> bool:
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_template_path(data_type: str) -> str | None:
    """返回指定类型的导入模板路径"""
    base_dir = os.path.join(current_app.static_folder, 'templates', 'excel')
    templates = {
        'person': os.path.join(base_dir, '人员导入模板.xlsx'),
        'building': os.path.join(base_dir, '建筑导入模板.xlsx'),  # 可扩展
    }
    return templates.get(data_type)


def export_data_to_excel(data_type: str, user) -> tuple[str, str]:
    """导出数据为 Excel 文件（支持人员/建筑，按用户网格权限过滤）"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exports_dir = current_app.config['EXPORTS_FOLDER']
    os.makedirs(exports_dir, exist_ok=True)

    user_grid_ids = get_user_grid_ids(user)

    if data_type == 'person':
        data = get_all_people_for_export(grid_ids=user_grid_ids)
        sheet_title = 'Person'
        filename_prefix = "人员数据"

        headers = [
            '姓名', '身份证号', '性别', '出生日期', '联系电话', '现住小区/建筑', '门牌地址', '所属网格',
            '人员类型', '是否重点人员', '重点类别', '户籍小区/建筑', '户籍地址', '户编号',
            '是否人户分离', '实际居住地', '是否已迁出', '迁出日期', '迁往地', '是否已死亡', '死亡日期',
            '民族', '政治面貌', '婚姻状况', '文化程度', '工作学习情况', '健康状况', '备注'
        ]
        comments = [
            '必填，真实姓名', '必填，18位身份证', '男/女', '格式：YYYYMMDD', '多个电话用;分隔',
            '系统内小区名称', '必填，如1单元101室', '自动关联', '常住人口/流动人口', '1=是，0=否',
            '多个类别用,分隔，如独居老人,低保户', '本社区内户籍小区', '外地户籍填写完整地址', '如有则填写',
            '1=是，0=否', '人户分离时填写实际居住地址', '1=是，0=否', '格式：YYYYMMDD', '迁往省市区',
            '1=是，0=否', '格式：YYYYMMDD', '如汉族、回族等', '如中共党员、群众等', '未婚/已婚/离异/丧偶',
            '小学/初中/高中/本科等', '在职/退休/无业等', '健康/良好/残疾/疾病等', '其他补充信息'
        ]

        # 单网格时文件名带网格名
        if user_grid_ids and len(user_grid_ids) == 1:
            grid = get_grid_by_id(user_grid_ids[0])
            grid_name = grid['name'] if grid else "未知网格"
            filename_prefix += f"_{grid_name}"

    elif data_type == 'building':
        data = get_all_buildings_for_export(grid_ids=user_grid_ids)
        sheet_title = 'Building'
        filename_prefix = "小区建筑数据"

        headers = [
            '小区/建筑名称', '类型', '所属网格', '详细地址', '建成年份', '户数', '楼栋数', '约居住人数',
            '企业数', '底商数', '是否通燃气', '物业费标准', '电梯数', '室内车位数', '室外车位数',
            '安全负责人', '负责人电话', '纬度', '经度', '开发商', '施工单位', '物业公司', '物业电话',
            '备注', '业委会联系人', '业委会电话', '业主姓名', '业主电话', '房东姓名', '房东电话', '商业类型'
        ]
        comments = [
            '必填，唯一名称', '住宅小区/私人住宅/大型出租房/商业建筑', '自动关联', '完整街道门牌号', '如2015',
            '总户数', '楼栋数量', '预估居住人数', '含居家办公企业', '底商店铺数量', '1=是，0=否',
            '如2.5元/㎡/月', '电梯总数', '室内停车位', '室外停车位', '安全负责人姓名', '联系电话',
            '可选，用于地图显示', '可选，用于地图显示', '开发商名称', '施工单位名称', '物业公司全称',
            '物业联系电话', '其他备注信息', '小区专用', '小区专用', '私人住宅专用', '私人住宅专用',
            '大型出租房专用', '大型出租房专用', '商业建筑专用（如商场、写字楼）'
        ]

    else:
        raise ValueError("不支持的导出类型")

    # 空数据时至少保留表头
    if not data:
        data = [{}]

    df = pd.DataFrame(data)

    filename = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = os.path.join(exports_dir, filename)

    # 使用 openpyxl 写入带注释的双表头 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # 第一行：正式表头
    ws.append(headers)
    # 第二行：注释说明
    ws.append(comments)
    # 第三行起：数据
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # 美化样式
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True)

    for cell in ws[1]:  # 表头加粗居中
        cell.font = bold_font
        cell.alignment = center_align

    for cell in ws[2]:  # 注释自动换行
        cell.alignment = wrap_align

    # 自动调整列宽（最大50）
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(file_path)
    logger.info(f"用户 {current_user.username} 导出 {data_type} 数据: {filename}")

    return file_path, filename


def process_import_excel(file, data_type: str, user) -> tuple[bool, str]:
    """处理 Excel 导入（支持人员，网格权限校验，自动清理临时文件）"""
    imports_folder = current_app.config['IMPORTS_FOLDER']
    temp_path = None

    try:
        if not allowed_file(file.filename):
            return False, '只支持 .xlsx 或 .xls 文件'

        filename = secure_filename(file.filename)
        temp_path = os.path.join(imports_folder, filename)
        file.save(temp_path)

        if data_type != 'person':
            return False, '暂不支持该类型导入'

        df = pd.read_excel(temp_path, dtype=str).fillna('')

        required_columns = ['姓名', '身份证号', '联系电话', '居住小区/建筑']
        if not all(col in df.columns for col in required_columns):
            return False, 'Excel 模板列不匹配，请下载最新模板'

        records = []
        fail_reasons = []

        for idx, row in df.iterrows():
            name = str(row['姓名']).strip()
            id_card = str(row['身份证号']).strip()
            phone = str(row['联系电话']).strip()
            building_name = str(row['居住小区/建筑']).strip()

            if not name or not id_card:
                fail_reasons.append(f"第 {idx+2} 行：姓名或身份证为空")
                continue

            building = get_building_by_name_or_address(building_name)
            if not building:
                fail_reasons.append(f"第 {idx+2} 行：未找到建筑 '{building_name}'")
                continue

            if not check_user_grid_permission(building['id']):
                fail_reasons.append(f"第 {idx+2} 行：无权操作该网格建筑 '{building_name}'")
                continue

            records.append({
                'name': name,
                'id_card': id_card,
                'phones': phone,
                'living_building_id': building['id'],
                'address_detail': str(row.get('门牌地址', '')),
                'notes': str(row.get('备注', ''))
            })

        success_count = bulk_insert_people(records) if records else 0
        fail_count = len(df) - success_count

        msg = f'导入完成：成功 {success_count} 条，失败 {fail_count} 条'
        if fail_reasons:
            msg += f"<br>失败原因示例：{fail_reasons[0]}（完整日志请查看服务器）"

        logger.info(f"用户 {user.username} 导入人员：成功 {success_count}，失败 {fail_count}")
        return True, msg

    except Exception as e:
        logger.error(f"用户 {user.username} 导入失败: {e}")
        return False, f'导入失败：{str(e)}'
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                logger.debug(f"清理临时导入文件: {temp_path}")
            except Exception as e:
                logger.warning(f"清理临时文件失败: {temp_path} - {e}")
