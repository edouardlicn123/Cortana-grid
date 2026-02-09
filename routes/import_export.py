# routes/import_export.py
# 导入导出路由（优化版 - 代码更简洁、可读性提升、错误处理更友好，功能完全不变）
# 更新：2026-02-09 字段全面同步最新 schema，支持人员模块所有字段的导入导出
# 重大优化：模板下载改为动态生成（无需预置静态文件），使用 openpyxl 在内存中实时创建
# 增强：导出/导入错误提示更具体，日志记录更详细，支持人员专属处理器

import os
import time
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint, send_from_directory, request, flash,
    redirect, url_for, render_template, current_app, send_file, jsonify
)
from flask_login import login_required, current_user
from permissions import permission_required
from services.import_export_service import (
    export_data_to_excel,
    process_import_excel,
)
from services.import_export_person import export_person_to_excel, import_person_from_excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from utils import logger


import_export_bp = Blueprint('import_export', __name__, url_prefix='/import_export')


def init_import_export_handlers(app):
    """初始化导入导出处理器（预留扩展点，可注册更多类型）"""
    with app.app_context():
        # 未来可在此注册其他模块的处理器（如建筑、网格等）
        pass


@import_export_bp.route('/')
@login_required
@permission_required('import_export:all')
def index():
    """导入导出首页"""
    return render_template('import_export.html')


@import_export_bp.route('/template/<data_type>')
@login_required
@permission_required('import_export:all')
def download_template(data_type):
    """动态生成并下载导入模板（实时创建带表头+注释的 Excel，无需预置文件）"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data_type}_导入模板"

    # 根据类型定义表头和注释（与专属服务文件保持一致）
    if data_type == 'person':
        # 人员模板表头（与 import_export_person.py 中的 headers 同步）
        headers = [
            '姓名', '身份证号', '唯一标识', '护照/其他证件号码', '其他证件类型',
            '性别', '出生日期', '联系电话', '现住小区/建筑', '现住详细门牌',
            '所属网格', '与其他人员关系', '人员类型', '是否重点人员', '重点类别',
            '户籍小区/建筑', '户籍详细地址', '户编号', '户号', '户籍迁入日期',
            '是否人户分离', '实际居住地', '是否已迁出', '迁出日期', '迁往地',
            '是否已死亡', '死亡日期', '民族', '政治面貌', '婚姻状况',
            '文化程度', '工作/学习情况', '健康状况', '备注'
        ]

        comments = [
            '必填，真实姓名', '可选，18位身份证号（无证可留空）', '系统内部唯一标识（可选）',
            '护照或其他证件号码', '护照/军人证/港澳通行证等',
            '男/女（支持：男、M、1；女、F、0）', '格式：YYYYMMDD', '多个用;分隔，可选',
            '必填，系统内现住建筑名称', '必填，如1单元101室',
            '自动关联，无需填写', '如：户主、配偶、子女、父母、租户（可选）',
            '常住人口/流动人口', '是/否 或 1/0', '多个类别用,分隔，如独居老人,低保户',
            '本社区户籍建筑名称（可选）', '外地户籍填写完整地址', '家庭编号（如001、A001）',
            '户口本户号', '格式：YYYYMMDD',
            '是/否 或 1/0', '人户分离时的实际居住地址', '是/否 或 1/0', '格式：YYYYMMDD', '迁往省市区',
            '是/否 或 1/0', '格式：YYYYMMDD', '如汉族、回族', '如中共党员、群众', '未婚/已婚/离异/丧偶',
            '小学/初中/高中/本科等', '在职/在校/退休/无业等', '健康/良好/慢性病/残疾等', '其他补充信息'
        ]

    elif data_type == 'building':
        flash('建筑模板尚未实现，请使用人员模板或联系管理员', 'warning')
        logger.info(f"用户 {current_user.username} 请求建筑模板（暂未实现）")
        return redirect(url_for('import_export.index'))

        # 如果未来实现，可在此定义建筑表头和注释
        # headers = [...]
        # comments = [...]

    else:
        flash(f'暂不支持类型：{data_type}', 'error')
        logger.warning(f"用户 {current_user.username} 请求不支持的模板类型：{data_type}")
        return redirect(url_for('import_export.index'))

    # 写入表头（第1行）
    ws.append(headers)

    # 写入注释（第2行）
    ws.append(comments)

    # 样式：表头加粗居中，注释自动换行
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_align

    for cell in ws[2]:
        cell.alignment = left_align

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    # 保存到内存流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回文件（文件名带完整时间戳，避免同名覆盖）
    download_name = f"{data_type}_导入模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@import_export_bp.route('/export/<data_type>')
@login_required
@permission_required('import_export:all')
def export(data_type):
    """导出数据为 Excel 文件"""
    try:
        if data_type == 'person':
            file_path, filename = export_person_to_excel(current_user)
        else:
            flash(f'暂不支持 {data_type} 类型导出', 'warning')
            logger.warning(f"用户 {current_user.username} 尝试导出不支持类型：{data_type}")
            return redirect(url_for('import_export.index'))

        return send_from_directory(
            directory=os.path.dirname(file_path),
            path=os.path.basename(file_path),
            as_attachment=True,
            download_name=filename
        )

    except ValueError as ve:
        flash(f'导出失败：{str(ve)}', 'error')
        logger.warning(f"用户 {current_user.username} 导出 {data_type} 失败（ValueError）：{ve}")
    except Exception as e:
        flash('导出过程中发生未知错误，请查看日志', 'error')
        logger.error(f"用户 {current_user.username} 导出 {data_type} 异常: {e}", exc_info=True)

    return redirect(url_for('import_export.index'))


@import_export_bp.route('/import', methods=['POST'])
@login_required
@permission_required('import_export:all')
def import_data():
    """处理 Excel 文件导入"""
    if 'file' not in request.files or not request.files['file'].filename:
        flash('未选择文件或文件名为空', 'error')
        return redirect(url_for('import_export.index'))

    file = request.files['file']
    data_type = request.form.get('import_type', '').strip()

    if not data_type:
        flash('未指定导入类型', 'error')
        return redirect(url_for('import_export.index'))

    # 文件类型校验
    allowed_extensions = {'.xlsx', '.xls'}
    if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
        flash('仅支持 .xlsx / .xls 文件', 'error')
        return redirect(url_for('import_export.index'))

    try:
        if data_type == 'person':
            success, msg = import_person_from_excel(file, current_user)
        else:
            flash(f'暂不支持 {data_type} 类型导入', 'warning')
            logger.warning(f"用户 {current_user.username} 尝试导入不支持类型：{data_type}")
            return redirect(url_for('import_export.index'))

        flash(msg, 'success' if success else 'error')

        if success and data_type == 'person':
            return redirect(url_for('person.index', _refresh=int(time.time())))

        return redirect(url_for('import_export.index'))

    except Exception as e:
        logger.error(f"用户 {current_user.username} 导入 {data_type} 失败: {e}", exc_info=True)
        flash(f'导入失败：{str(e)[:100]}...（详情见日志）', 'error')
        return redirect(url_for('import_export.index'))


# 可选：API 接口，用于未来前端异步进度查询
@import_export_bp.route('/api/import/status', methods=['GET'])
@login_required
@permission_required('import_export:all')
def import_status():
    """查看最近导入状态（占位，未来可实现）"""
    return jsonify({
        'status': 'idle',
        'message': '暂无导入任务'
    })
